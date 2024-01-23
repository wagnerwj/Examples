from openpyxl import load_workbook
import csv
import numbers


headerrow=["code", "modifier", "workRvu", "nonPeRvu", "peRvu", "mpeRv", "description", "wcFactor", "baseUnits"]


cpt=load_workbook('NAME')
anes=load_workbook('NAME')
aneswc=anes.active
cptws=cpt.active
with open('csvout.csv','w') as newCSV:
  zipWriter=csv.writer(newCSV)
  zipWriter.writerow(headerrow)
  for row in cptws.iter_rows(min_row=13, max_col=11,values_only=True):

    if row[2]!='D':
      result=[]
      result.append(row[0])
      result.append(row[1])
      for index in range(3,7):
        if isinstance(row[index],number.Number):
          result.append(row[index])
        else:
          result.append(0)
      if row[10]!=None:
        result.append(row[10].strip())
      else:
        result.append(" ")
      result.append(row[8])
      result.append(0)

      code=row[0]
      printed=False
      for anrow in aneswc.iter_rows(max_row=288, min_row=13, max_col=3):
        if anrow[0].value == code:
          result[8]=anrow[2].value
          for mod in modlist:
            result[1]=mod
            zipWriter.writerow(result
            printed=True
          break

      if not printed:
            zipWriter.writerow(result)
